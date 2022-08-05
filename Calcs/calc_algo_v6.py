# %%
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import datetime
import functools
import sys

# %%
#import warnings
#warnings.filterwarnings('error',category=FutureWarning)
#warnings.filterwarnings('error',category=UserWarning)

# %%
def output_result_to_excel(file_name):
    '''
    helper function to write result dataframe to excel file, roughly in following shape

    ---------------------------------------------------------------------------------
    ideal code
    ---------------------------------------------------------------------------------
    case name
    ---------------------------------------------------------------------------------
    top1 code name
    top1 code mark
    ......
    top5 code name
    top5 code mark
    ---------------------------------------------------------------------------------
    codeX mark
    ...
    ---------------------------------------------------------------------------------
    
    sheet2 will contain the summary of result.

    '''
    # %%
    # those value is for test the output format
    '''df = pd.DataFrame({'case1':{'code1':1.0,'code2':0.9,'code3':0.8,'code4':0.7,'code5':0.9},\
                    'case2':{'code1':0.3,'code2':0.7,'code3':0.6,'code4':0.7,'code5':0.9} })
    file_name = 'test_.xlsx'
    ideal_code = pd.DataFrame({'case1':{'Ideal code':'code1'},\
                    'case2':{'Ideal code':'code2'} })
                    '''
    global calc_results
    df = calc_results['df_results_adjust']
    assert(not df.empty)
    ideal_code = parsed_data['ideal_code_for_case']
    index_result = calc_results['indices_result']
    score_result = calc_results['scores_result']
    top_5_codes = calc_results['df_top5_codes']
    top_5_value = calc_results['df_top5_mark']

    # %%
    wb = Workbook()
    ws = wb.active

    # Ideal code to be as the top line.  skip header row  
    # because after header row has index name, need skip
    rows = dataframe_to_rows(ideal_code,index=True, header=False)
    next(rows)
    for row in rows:
        if row == [None]:
            continue
        ws.append(row)

    rows = dataframe_to_rows(df,index=True,header=True)
    # add first row (which is header)
    first_row = next(rows)
    ws.append(first_row)

    # %%
    all_cases = config['cases_list']
    top_5_code_rows = dataframe_to_rows(top_5_codes,index=True,header=False)
    top_5_value_rows = dataframe_to_rows(top_5_value,index=True,header=False)

    # after header row there always one row of [None], which normally should be for index name,
    # need to skip that row.
    for code_row in top_5_code_rows:
        value_row = next(top_5_value_rows)
        if code_row == [None]:
            continue
        ws.append(code_row)
        ws.append(value_row)

    # output the index and score row.
    index_result = calc_results['indices_result']
    index_rows = dataframe_to_rows(index_result,index=True,header=False)
    for index_row in index_rows:
        if index_row == [None]:
            continue
        ws.append(index_row)

    score_result = calc_results['scores_result']
    score_rows = dataframe_to_rows(score_result,index=True,header=False)
    for score_row in score_rows:
        if score_row == [None]:
            continue
        ws.append(score_row)

    # add the rest rows from mark calculation dataframe for all codes
    # %%
    for row in rows:
        if row == [None]:  # skip the index name row.
            continue
        ws.append(row)

    # format style Pandas will make font to bold, like table header
    # left column and top row (ideal code row) and 2nd row (case name row)
    style_cells = ws['A']
    # top 5 code name and value, total 10 rows
    for i in range(1,15):    
        style_cells += ws[i]
    for cell in style_cells:
        cell.style = 'Pandas'

    # format the top 5 row number value cells into percentage
    for i,ws_row in enumerate(ws.iter_rows(min_row=4, max_row=12, min_col=2,max_col=ws.max_column)):
        if i % 2 == 0:
            for cell in ws_row:
                cell.number_format = '0.00%'
    # format all calculated code value number cells into percentage
    for ws_row in ws.iter_rows(min_row=15, max_row=ws.max_row, min_col=2,max_col=ws.max_column):
        for cell in ws_row:
            cell.number_format = '0.00%'

    for col in range(2,ws.max_column+1):
        cell = ws.cell(3, col)
        # format top 1 row to green color
        cell.fill = PatternFill(start_color="378805", end_color="378805", fill_type="solid")
        # format the ideal code row, match top 1 be grey, not match be red.
        cell2 = ws.cell(1, col)
        if cell2.value == cell.value:
            cell2.fill = PatternFill(start_color="808088", end_color="808088", fill_type="solid")
        else:
            cell2.fill = PatternFill(start_color="bb0a1e", end_color="bb0a1e", fill_type="solid")

        # format the index and score row to yellow color.
        cell3 = ws.cell(13, col)
        cell4 = ws.cell(14, col)
        cell3.fill = PatternFill(start_color="ffff00", end_color="ffff00", fill_type="solid")
        cell4.fill = PatternFill(start_color="ffff00", end_color="ffff00", fill_type="solid")

    # make a summary infomation into a dataframe.
    # then create and output summary sheet (sheet2)
    # %%
    ind = pd.Index(range(1,7),name='Sl No')
    total_cases = ws.max_column-1
    counts = [0 for _ in range(5)]
    percent = [0.0 for _ in range(5)]
    for case in all_cases:
        for i in range(5):
            if top_5_codes.loc[f'Top {i+1}', case] == ideal_code.loc['Ideal code',case]:
                counts[i] += 1
    for i in range(5):
        percent[i] = counts[i]/total_cases
    position = ['1st', '2nd', '3rd', '4th', '5th']
    data = np.array([('Total # of test cases', total_cases, np.nan)]+\
           [(f'Total # of cases containing ideal code in {position[i]} position', counts[i], percent[i]) for i in range(5)],\
           dtype=[('Description','U100'),('No','i4'),('%', 'f4')])
    df = pd.DataFrame(data,index=ind)
    
    # %%
    ws2 = wb.create_sheet('sheet2')
    rows = dataframe_to_rows(df,index=True,header=True)
    for row in rows:
        if row == [None]:  # skip the index name row.
            continue
        ws2.append(row)
    style_cells = ws2['A'] + ws2[1]
    for cell in style_cells:
        cell.style = 'Pandas'
    style_cells = ws2[get_column_letter(ws2.max_column)]
    for cell in style_cells:
        if isinstance(cell.value,float):
            cell.number_format = '0.00%'

    wb.save(file_name)

    return (0, 'Ok')


# %%
def process_master_file(master_file):
    '''
    function to be called to initialize process of master file and adjust file.
    arguments:
        master_file: the file name of master sheet 
                      (include path to be able to open it)
        
    return value:
        (1, 'Error Message'): master sheet has something wrong cannot proceed
        (3, 'Error Message): case sheet has something wrong cannot proceed
    
    process result:
        the processed data will be kept in dictionary:
        parsed_data, config
    '''
    global config, parsed_data

    # %%
    master_empty_coef, _, _ = config['cell_value_settings']
    index_col_name, score_col_name = config['special_columns']

    # %%
    #master_file = '../Sample master sheet - v2 - 30.01.2022.xlsx'
    try: 
        df_master = pd.read_excel(master_file, 
                        header=0, # the first row of master file is header 
                        index_col=(1,2,3,4) # the second, fifth columns (cat, cat type, feature, subfeature) as index
                    ).drop('Unnamed: 0', axis=1) # 'Unnamed: 0' is the index column in excel which no need 
    except FileNotFoundError:
        #print(1, f'{master_file} not found')
        return (1, f'Cannot find {master_file}')

    # %%
    # remove the empty line, which will be duplicated index and all NaN value (second line)
    df_master = df_master[~ df_master.index.duplicated(keep='first')]


    # %%
    master_index = df_master.index
    master_cols = df_master.columns
    
    # %%
    config['match_cat'] = master_index[master_index.get_level_values('Cat. Type')==1]\
                             .get_level_values('Category').unique()
    config['codes_list'] = master_cols[2:]
    config['adjust_cat'] = master_index[master_index.get_level_values('Cat. Type')==2]\
                             .get_level_values('Category').unique()

    
    # %%
    # get the matching part from master sheet 
    df_mark_master = df_master.loc[config['match_cat'], config['codes_list']]

    # %%
    # get the index and score column (value for each (feature subfeature) combinations)
    s_index_master = df_master[index_col_name]\
                        .loc[config['match_cat']]\
                        .droplevel((0,1),axis=0)
    s_score_master = df_master[score_col_name]\
                        .loc[config['match_cat']]\
                        .droplevel((0,1),axis=0)



    # %%
    def master_cell_map(x):
        '''extra the value of each (feature, subfeature) combination for each code.
        value is the last part of cell content
        '''
        if pd.isnull(x): return master_empty_coef
        else :
            # split cell content and use last part to extract the value
            x = x.split(',')
            return float(x[-1])
  
    # maps sheet cell content to numeric value.
    # level 1 index is cat. type which no need anymore.
    df_mark_master = df_mark_master.applymap(master_cell_map).droplevel((1,), axis=0)


    # %%
    # calculate the possible highest benchmark of each code (column) each cat. (row)
    # for each feature take the maximum from all subfeatures.
    # then add up score of all features (within one category)
    cat_maximum_score = df_mark_master.groupby(level=['Category', 'Feature Code'])\
                        .max()\
                        .groupby(level=['Category'])\
                        .sum()


    # %%
    # dataframe index now is (cat., feature, subfeature) all combinations
    # drop level 0 cat. then remain is (Feature, subfeature), should be unique
    df_mark_master = df_mark_master.droplevel((0,),axis=0)
    match_feature_items = df_mark_master.index  
    if not match_feature_items.is_unique:
        #print(1, '(Feature, subfeature) not unique')
        return (1, '(Feature, subfeature) not unique')


    # %%
    parsed_data['match_feature_items'] = match_feature_items
    parsed_data['df_mark_master'] = df_mark_master
    parsed_data['cat_maximum_score'] = cat_maximum_score
    parsed_data['s_index_master'] = s_index_master
    parsed_data['s_score_master'] = s_score_master

    # %%
    # get the ratio adjusting part from master sheet 
    df_adjust_probability = df_master.loc[config['adjust_cat'], config['codes_list']]\
                            .droplevel((0,1), axis=0)
    # map NaN value to 1 if have any empty cell. (no adjustment)
    df_adjust_probability.fillna(1, inplace=True)


    # %%    
    adjust_feature_items = df_adjust_probability.index
    if not adjust_feature_items.is_unique:
        #print(1, '(Feature, subfeature) not unique')
        return (1, '(Feature, subfeature) not unique')

    # %%
    parsed_data['df_adjust_probability'] = df_adjust_probability
    parsed_data['adjust_feature_items'] = adjust_feature_items


    # %%
    # extrace the age span for adjustment calculation.
    age_span = adjust_feature_items[adjust_feature_items.get_level_values(0)=='Age']\
                .droplevel('Feature Code')\
                .to_series()\
                .apply(lambda x: list(map(int, x.split('-'))))
    age_span = pd.DataFrame(age_span.to_list(), index=age_span.index)
    parsed_data['age_span'] = age_span


    # %%
    return (0, 'Success')


def process_age_adjust():
    '''
    according to each case data of birth, update the age adjustment value of the case
    '''
    global config, parsed_data

    # %%
    dob_feature_name = config['dob_feature'][0]
    df_cat_special_case = parsed_data['df_cat_special_case']
    all_cases = config['cases_list']
    age_span = parsed_data['age_span']


    # %%
    assert(dob_feature_name)
    assert(not df_cat_special_case.empty)

    if dob_feature_name not in df_cat_special_case.index.get_level_values(0):
        #print(3, 'Error in getting Date of Birth')
        return(3, 'Error in getting Date of Birth')

    # %%
    df_adjust_age_TF = pd.DataFrame([[False]*len(all_cases)]*len(age_span.index),
                index=pd.MultiIndex.from_product([['Age'], age_span.index]),
                columns=all_cases)


    # %%
    for case in all_cases:
        # for each case, try to calculate age according to dob
        dob = df_cat_special_case.loc[(dob_feature_name,slice(None)), case].squeeze()
        if pd.notnull(dob):
            try: 
                dob = datetime.datetime.strptime(dob, '%d/%m/%Y')
            except (TypeError, ValueError):
                return (3, f'Error in process Date of Birth {dob} with type {type(dob)}')

            age = int((datetime.datetime.now() - dob).days/365)
            # if we can calculate the age, then try to match it into one age_span
            # and set the corresponding adjustment data to True.
            # for later doing adjustment calculations. 
            df_adjust_age_TF.loc[\
                ('Age',age_span.index[(age_span[0]<=age) & (age<=age_span[1])]),\
                case ]\
                = True
    
    # %%
    df_adjust_case_TF = pd.concat([parsed_data['df_adjust_case_TF'], df_adjust_age_TF], 
                              axis=0)
    
    parsed_data['df_adjust_case_TF'] = df_adjust_case_TF
    
    # %%

    return (0, 'Ok')


def calc_special_column1():
    '''
    calculate index value
    '''
    global config, parsed_data, calc_results

    # %%
    index_master = parsed_data['s_index_master']
    all_cases = config['cases_list']
    df_mark_case_selected = parsed_data['df_case_mark_cats_TF']

    df_index_result = pd.DataFrame(columns=all_cases, index=['Index'], dtype=np.float64)
    for case in all_cases:
        df_index_result.loc['Index',case] = index_master[df_mark_case_selected[case]].sum()
    
    # %%
    calc_results['indices_result'] = df_index_result


    # %%
    return (0, 'Ok')

def calc_special_column2():
    '''
    calculate score value
    '''
    global config, parsed_data, calc_results

    # %%
    score_master = parsed_data['s_score_master']
    all_cases = config['cases_list']
    df_mark_case_selected = parsed_data['df_case_mark_cats_TF']

    df_score_result = pd.DataFrame(columns=all_cases, index=['Score'], dtype=np.unicode_)
    for case in all_cases:
        case_score = score_master[df_mark_case_selected[case]]\
                        .to_numpy()
        
        if 'A' in case_score:
            score = 'A'
        elif 'B' in case_score:
            score = 'B'
        else:
            score = np.nan
        df_score_result.loc['Score', case] = score

    # %%
    calc_results['scores_result'] = df_score_result

    # %%
    return (0, 'Ok')

# %%
def mark_calculation():
    '''
    perform mark calculation for all cases with all codes.
    '''
    global config, parsed_data, calc_results
    # %%
    code_names = config['codes_list']
    all_cases = config['cases_list']

    df_mark_case = parsed_data['df_mark_case']
    df_mark_master = parsed_data['df_mark_master']
    df_case_count_subfeature = parsed_data['df_case_count_subfeatures']
    cat_maximum_score = parsed_data['cat_maximum_score']
    df_case_cat_attempted = parsed_data['df_case_attempted_cats']

    # %%
    # for each case, and each code, calculate one score
    # if one case for one feature choose multiple subfeatures, then use average
    # create an empty dataframe first, then fill it with all the calculation values
    df_results = pd.DataFrame(index=code_names, \
                    columns=all_cases,\
                    dtype=np.float64)

    # %%
    for case in all_cases:
        for code in code_names:
            # vector itemwise multiply: case choosen (feature, subfeature), with master values for (feature, subfeature)
            # then sum over all subfeatures 
            score = (df_mark_master[code] * df_mark_case[case] )\
                            .groupby(level=['Feature Code']).sum()
            # if this case for one feature has multiple subfeatures, need divide by the counts
            # this do the vector itemwise divide
            score = score/df_case_count_subfeature[case]
            # sum over all features, get the final score value.
            df_results.loc[code,case] = score.sum()
    

    # %%
    # use each case score to divide maximum possible score for the code, get percentage
    # the result score divide by respective code maximum, get the percentage value
    case_maximum_score = (cat_maximum_score.T).dot(df_case_cat_attempted)

    # calculate the percentage using divide of result v.s. maximum
    df_results_percent = df_results.divide(case_maximum_score)
    
    # %% 
    # there should not be any non-zero value/0. (which will results in inf)
    # if cell in case_maximum_score is 0,
    # then the corresponding df_results should also be 0
    assert(False == np.isinf(df_results_percent).any().any())

    # %%
    # due to some cell is 0/0 and result is NaN, for those cell map to 0 value.
    df_results_percent.fillna(0, inplace=True)
    
    # store the calculation reault into calc_results dict.
    calc_results['df_results_percent_non_adjust'] = df_results_percent

    # %%
    return (0, 'Ok')


def adjustment_calculation():
    '''
    do adjustment / fine-tune
    '''   
    global config, parsed_data, calc_results
    # %%
    # stage 1 calculation must already done.
    # adjustmeent table must already been processed.
    df_results_percent = calc_results['df_results_percent_non_adjust']
    df_adjust_probability = parsed_data['df_adjust_probability']
    assert(not df_results_percent.empty)
    assert(not df_adjust_probability.empty)

    # start from the percent result to do adjust
    df_results_adjust = df_results_percent.copy()

    all_cases = config['cases_list']
    df_adjust_case_TF = parsed_data['df_adjust_case_TF']

    # %%
    for case in all_cases:
        # for each case, if adjust (feature subfeature) is not null,
        # then need make adjustment accordingly with the probability ratio table.
        idx = df_adjust_case_TF.index[df_adjust_case_TF[case]]
        if not idx.empty:  # has some adjustment for this case
            if idx.size == 1:
                df_adjust_ratio = (df_adjust_probability.loc[idx,:]).squeeze()
            else:
                df_adjust_ratio = functools.reduce(\
                    lambda x1, x2: (0, x1[1]*x2[1]),\
                    (df_adjust_probability.loc[idx,:]).iterrows())
                df_adjust_ratio = df_adjust_ratio[1]

            df_results_adjust[case] *= df_adjust_ratio

    calc_results['df_results_adjust'] = df_results_adjust

    # %%
    calc_results['df_results_percent_non_adjust']
    # %%
    return (0, 'Ok')

def find_top5():
    # %%
    # add top 5 code row (which gives the best top 5 code for each case)
    global calc_results
    df = calc_results['df_results_adjust']
    assert(not df.empty)

    # %%
    all_cases = config['cases_list']
    index = [f'Top {i}' for i in range(1,6)]
    # create empty dataframe for top 5 code and top 5 value.
    top_5_value = pd.DataFrame(columns=all_cases, index=index, dtype=np.float64)
    top_5_code = pd.DataFrame(columns=all_cases, index=index, dtype=np.unicode_)
    for col in all_cases:
        df_order_value = df[col].sort_values(axis=0,ascending=False)\
                            .head(5)
        top_5_value.loc[:,col] = df_order_value.to_numpy()
        top_5_code.loc[:,col] = df_order_value.index
    
    # %%
    calc_results['df_top5_codes'] = top_5_code
    calc_results['df_top5_mark'] = top_5_value

    # %%
    return (0, 'Ok')

# %%
def process_case_file(case_file):
    '''
    read case sheet file. extrace the data for all cases. data stored in parsed_data.
    then call the calculation process to do calculation.
    calculation results stored in calc_results dict.
    '''
    global config, parsed_data
    # %%
    _, case_empty_coef, case_filled_coef = config['cell_value_settings']

    # %%
    #case_file = '../Sample test sheet - v2 - 30.01.2022.xlsx'
    try:
        df_case = pd.read_excel(case_file, 
                    header=0, # the first row of master file is header 
                    index_col=(1,2,3) # the 2nd, 3rd and 4th column (cat, feature, subfeature) as index
                    ).drop('Sl No',axis=1) # 'Sl No' is the column of index in excel which no need.
    except FileNotFoundError:
        #print(3, f'{case_file} cannot be found')
        return (3, f'{case_file} cannot be found')


    # %%
    # get all the cases from case sheet
    config['cases_list'] = df_case.columns

    # %%
    # extract the 'Ideal code' list for different cases
    ideal_code_for_case = df_case.loc[(slice(None),'Ideal code',slice(None)),:]\
                          .droplevel([0,2],axis=0)
    parsed_data['ideal_code_for_case'] = ideal_code_for_case

    # %%
    # remove those non-catogory rows. 
    df_case = df_case[~ df_case.index.get_level_values(0).isna()]

    # %%
    assert(not config['match_cat'].empty)
    assert(not config['adjust_cat'].empty)
    assert(config['special_cat'])
    # the C0 special category, which has date of birth information.
    special_categories = config['special_cat']
    # the categories for matching and mark calculation
    match_categories = config['match_cat']
    # the categories for adjustment / fine-tune.
    # not include age_span which not appear in test sheet
    adjust_categories = config['adjust_cat'].drop(config['age_span_cat'][0])

    # separate case sheet infor according to categories
    df_cat_special_case = df_case.loc[special_categories, :].droplevel('Category')
    df_mark_case = df_case.loc[match_categories, :]
    df_adjust_case = df_case.loc[adjust_categories, :]
    

    # %%
    # extract the categories attempted for each cases
    df_case_cat_attempted = df_mark_case.groupby(level=['Category'])\
                                .apply(lambda x: (pd.notnull(x)).any(axis=0))
    
    # %%
    df_mark_case = df_mark_case.droplevel('Category',axis=0)
    df_adjust_case = df_adjust_case.droplevel('Category',axis=0)

    # %%
    # check if the (feature, subfeature) from master sheet match with the case sheet
    if not (parsed_data['match_feature_items'].sort_values() == df_mark_case.index.sort_values()).all():
        #print(3, 'Match categories features in master and case not same.')
        return (3, 'Match categories features in master and case not same.')


    # %%
    # check if the (feature, subfeature) from probability ratio table match with the case sheet
    adjust_feature_items = parsed_data['adjust_feature_items']
    if not (adjust_feature_items.drop('Age', level=0).sort_values() == df_adjust_case.index.sort_values()).all():
        #print(3,'Ratio adjust categories features in master and case not same.')
        return (3,'Ratio adjust categories features in master and case not same.')


    # %%
    df_case_mark_cats_TF = df_mark_case.notnull()

    # calculate within each feature, how many subfeatures selected
    # most of case will only choose one subfeature in each feature.
    # but there are some cases may choose multiple subfeatures in one feature.
    # %%
    df_case_subfeature_counts = df_case_mark_cats_TF.groupby(level=[0,]) \
                                .sum()


    # conver case infomation table into numeric for calculation
    # %%
    df_mark_case = df_mark_case.applymap(lambda x: case_empty_coef if pd.isnull(x)\
                                         else case_filled_coef)
    
    # conver adjust feature of case to (True False) value
    # %%
    df_adjust_case_TF = df_adjust_case.notnull()
    
   
    parsed_data['df_mark_case'] = df_mark_case
    parsed_data['df_case_mark_cats_TF'] = df_case_mark_cats_TF
    parsed_data['df_adjust_case_TF'] = df_adjust_case_TF

    parsed_data['df_cat_special_case'] = df_cat_special_case

    parsed_data['df_case_attempted_cats'] = df_case_cat_attempted
    parsed_data['df_case_count_subfeatures'] = df_case_subfeature_counts

    # %%
    # according to date of birth, find the age adjustment ratio
    result = process_age_adjust()
    if result[0]:
        return result
    
    return (0, 'Success')

def do_calculation():

    mark_calculation()
    adjustment_calculation()
    find_top5()

    calc_special_column1()
    calc_special_column2()

    return (0, 'Ok')


# %%
# some configuration information
config = {}
# tuple of value settings for different type of cells
# position 0, master sheet empty cell value. defaut 0
# position 1, case sheet empty cell value. default 0
# position 2, case sheet filled cell value. default 1
config['cell_value_settings'] = (0,0,1)
# codes, will read from master sheet, from column H and after.
config['codes_list'] = None
# categories to be used to matching and calculate marks
# will read from master sheet, type 1 categories
config['match_cat'] = None
# categories to be used to look up probability and adjust fine tune mark
# will read from master sheet, type 2 categories
config['adjust_cat'] = None
# age specail category. the feature name is 'age'. 
# subfeature will has age span information needs to parse
config['age_span_cat'] = ['C1000'] 
# special cat, special feature, not used for normal calculation.
# only DATE OF BIRTH, will be used to calculate age.
config['special_cat'] = ['C0']
config['dob_feature'] = ['Feature 3']
# special requirement regarding the Index and Score columns, the calculation will be
# based on the corresponding (feature, subfeature) is choosen by the case or not.
config['special_columns'] = ['Index', 'Score']
# list of all cases. read from case file.
config['cases_list'] = None


# %%
# after parse the original master sheet, and case sheet,
# the data from those sheet will be kept in the parsed_data
parsed_data = {}
# to keep mark calculation table from master sheet.
parsed_data['df_mark_master'] = None 
# to keep maximum possible benchmark for each categories.
parsed_data['cat_maximum_score'] = None 
# to keep possibillity table .
parsed_data['df_adjust_probability'] = None 
# to keep the index special value 
# it is a column in master sheet, one value for each (feature, subfeature) combination
# empty may exist means no value for those (feature, subfeature)
parsed_data['s_index_master'] = None 
# to keep the score special value
# it is a column in master sheet, one value for each (feature, subfeature) combination
# empty may exist means no value for those (feature, subfeature)
parsed_data['s_score_master'] = None 
# to keep all (feature, subfeature) for mark calculation. pandas index object
parsed_data['match_feature_items'] = None 
# to keep all (feature, subfeature) for adjust probability. pandas index object
parsed_data['adjust_feature_items'] = None 
# to keep the age spans for adjustment look up
parsed_data['age_span'] = None 


# %%
# to keep all the ideal code infomation for different cases
parsed_data['ideal_code_for_case'] = None

# to keep special category (c0) related information from case sheet
# this category has DATA OF BIRTH which needs to be used for age calculation
parsed_data['df_cat_special_case'] = None 

# to keep all the mark calculation information from case sheet
parsed_data['df_mark_case'] = None
# to keep all adjust table information from all cases 
parsed_data['df_adjust_case_TF'] = None 

# record each case how many categories attempted
parsed_data['df_case_attempted_cats'] = None 

# record each feature, each case choose how many subfeatures. 
# few case may multiple choice of more than one subfeatures from one feature
parsed_data['df_case_subfeature_counts'] = None
# a True/False value to show test case choosen or not choosen (feature, subfeature)
parsed_data['df_case_mark_cats_TF'] = None


# %%
# to keep the calculation results
calc_results = {}
# record the mark calculation result, before adjustment
calc_results['df_results_percent_non_adjust'] = None 
# record the mark after adjust/fine tune.
calc_results['df_results_adjust'] = None 
# record the top5 codes name and mark for each case
calc_results['df_top5_codes'] = None
calc_results['df_top5_mark'] = None
# record the subfeature_values (now have index and score) for each cases
calc_results['indices_result'] = None
calc_results['scores_result'] = None

# %%
if __name__ == '__main__':
    # %%
    import time
    master_file = sys.argv[1]
    case_file = sys.argv[2]
    result = process_master_file(master_file)
    if result[0]:
        print(result)
        assert(0)

    t0 = time.perf_counter()

    result = process_case_file(case_file)
    if result[0]:
        print(result)
        assert(0)

    t1 = time.perf_counter()

    do_calculation()

    t2 = time.perf_counter()

    output_file = case_file.split('.')[-2].split('\\')[-1]+'_result_'+\
        datetime.datetime.now().strftime('%d%m%y')+'.xlsx'
    output_result_to_excel(output_file)
    t3 = time.perf_counter()

    print(f'finished. read and process case file time : {t1-t0}')
    print(f'          calculation time: {t2-t1}')
    print(f'          time for saving file : {t3-t2}')
